import os
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, Menu
from PIL import Image, ImageTk
import openpyxl


class ImageAnnotator:
    def __init__(self, root):
        self.root = root
        self.root.title("SpeedyBatv2.0")

        # Open app in full screen
        self.root.state('zoomed')  # For Windows

        self.folder_path = ""
        self.image_list = []
        self.image_index = 0
        self.annotations = {}
        self.fields = []  # List to store added field names
        self.field_vars = {}  # Dictionary to map field names to IntVar for checkboxes
        self.field_shortcuts = {}  # Dictionary to map field names to shortcut keys
        self.xlsx_file = ""  # Path to annotations.xlsx
        self.ws = None  # Worksheet object for writing annotations
        self.wb = None  # Workbook object

        # Menu
        self.menu = Menu(self.root)
        self.root.config(menu=self.menu)

        self.fields_menu = Menu(self.menu)
        self.menu.add_cascade(label="Add Fields", menu=self.fields_menu)
        self.fields_menu.add_command(label="Add New Field", command=self.add_field)

        # Toolbar
        self.toolbar = tk.Frame(self.root)
        self.toolbar.pack(side=tk.TOP, fill=tk.X)

        self.load_button = tk.Button(self.toolbar, text="Load Folder", command=self.load_folder)
        self.load_button.pack(side=tk.LEFT)

        self.backward_button = tk.Button(self.toolbar, text="<", command=self.show_previous_image)
        self.backward_button.pack(side=tk.LEFT)

        self.forward_button = tk.Button(self.toolbar, text=">", command=self.show_next_image)
        self.forward_button.pack(side=tk.LEFT)

        # Image display
        self.image_label = tk.Label(self.root)
        self.image_label.pack()

        self.checkboxes_frame = tk.Frame(self.root)
        self.checkboxes_frame.pack()

        # Bind key press events for shortcuts
        self.root.bind("<Key>", self.toggle_field)
        self.root.bind("<Left>", lambda event: self.show_previous_image())  # Left arrow for previous
        self.root.bind("<Right>", lambda event: self.show_next_image())  # Right arrow for next

    def load_folder(self):
        self.folder_path = filedialog.askdirectory()
        if not self.folder_path:
            return

        self.image_list = [f for f in os.listdir(self.folder_path) if
                           f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]
        if not self.image_list:
            messagebox.showerror("Error", "No images found in the selected folder.")
            return

        # Initialize the path for the Excel file
        self.xlsx_file = os.path.join(self.folder_path, 'annotations.xlsx')

        # Check if any fields have been added
        if not self.fields:  # Proceed only if no fields have been added
            if os.path.exists(self.xlsx_file):
                self.read_existing_annotations()
            else:
                # Initialize annotations for images based on fields
                self.annotations = {image: {field: None for field in self.fields} for image in self.image_list}
                self.create_annotation_file()  # Call without arguments
                self.image_index = 0  # Start at the first image if there's no existing file
        else:
            # If fields have been added, initialize annotations without reading the existing file
            self.annotations = {image: {field: None for field in self.fields} for image in self.image_list}
            self.create_annotation_file()  # Call without arguments
            self.image_index = 0  # Start at the first image

        self.show_image()

    def read_existing_annotations(self):
        self.wb = openpyxl.load_workbook(self.xlsx_file)
        self.ws = self.wb.active

        # Read field names from the header row (assuming the first row contains field names)
        headers = [cell.value for cell in self.ws[1]]
        self.fields = headers[1:]  # Ignore the "Image" column

        # Ask for shortcut assignments for each field
        for field in self.fields:
            shortcut = simpledialog.askstring("Assign Shortcut", f"Assign a shortcut for '{field}':")
            if shortcut and len(shortcut) == 1:
                self.field_shortcuts[field] = shortcut

        # Rebuild annotations dictionary based on existing data
        self.annotations = {}
        last_annotated_image = None  # To track the last annotated image

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            image = row[0]
            field_values = row[1:]
            self.annotations[image] = {self.fields[i]: ('x' if field_values[i] == 'x' else None) for i in
                                       range(len(self.fields))}

            # Check if the current image has any annotations
            if any(value == 'x' for value in field_values):
                last_annotated_image = image  # Update the last annotated image
                print(f"Annotated Image Found: {image}")  # Debugging line

        print(f"Last Annotated Image: {last_annotated_image}")  # Debugging line

        # Set the image index to the last annotated image if found
        if last_annotated_image and last_annotated_image in self.image_list:
            self.image_index = self.image_list.index(last_annotated_image) + 1
            print(
                f"Starting at Last Annotated Image: {last_annotated_image} (Index: {self.image_index})")  # Debugging line
        else:
            self.image_index = 0  # Fallback to first image
            print(
                f"No last annotated image found. Starting at First Unannotated Image (Index: {self.image_index})")  # Debugging line

    def get_first_unannotated_index(self):
        for index, image in enumerate(self.image_list):
            # Check if all fields for the image are annotated
            if any(value is None for value in self.annotations.get(image, {}).values()):
                return index
        return 0  # If all images are annotated, start at the first image

    def create_annotation_file(self):
        if os.path.exists(self.xlsx_file):
            # Ask for confirmation to overwrite
            if not messagebox.askyesno("Warning", "The annotation file already exists. Do you want to overwrite it?"):
                return  # Exit if user chooses not to overwrite

            os.remove(self.xlsx_file)  # Remove the existing file if the user wants to overwrite

        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        headers = ["Image"] + self.fields
        self.ws.append(headers)

        # Populate the "Image" column with all filenames
        for image in self.image_list:
            row = [image] + [None] * len(self.fields)  # Fill with None for fields
            self.ws.append(row)

        self.wb.save(self.xlsx_file)

    def show_image(self):
        if self.image_index < 0 or self.image_index >= len(self.image_list):
            return

        image_path = os.path.join(self.folder_path, self.image_list[self.image_index])
        image = Image.open(image_path)
        w, h = self.root.winfo_width(), self.root.winfo_height()  # Get window width and height
        image.thumbnail((w, h))  # Resize the image to fit the window
        photo = ImageTk.PhotoImage(image)
        self.image_label.configure(image=photo)
        self.image_label.image = photo

        # Change title to image name
        self.root.title(os.path.basename(image_path))
        self.update_checkboxes()

    def update_checkboxes(self):
        # Clear previous checkboxes
        for widget in self.checkboxes_frame.winfo_children():
            widget.destroy()

        # Show all fields as checkboxes
        for field in self.fields:
            var = tk.IntVar()
            self.field_vars[field] = var  # Store IntVar for the field
            checkbox = tk.Checkbutton(self.checkboxes_frame, text=field, variable=var,
                                      command=lambda field=field, var=var: self.update_annotation(field, var))
            checkbox.pack(anchor='w')

            # Restore state of checkbox if it exists in annotations
            if self.image_list and self.image_index < len(self.image_list):
                if self.annotations.get(self.image_list[self.image_index], {}).get(field):
                    var.set(1)

    def update_annotation(self, field, var):
        checked = var.get()
        self.annotations[self.image_list[self.image_index]][field] = 'x' if checked else None

    def show_next_image(self):
        if self.image_index < len(self.image_list) - 1:
            self.save_annotations()  # Save current annotations before moving
            self.image_index += 1
            self.show_image()

    def show_previous_image(self):
        if self.image_index > 0:
            self.save_annotations()  # Save current annotations before moving
            self.image_index -= 1
            self.show_image()

    def save_annotations(self):
        # Get the row index for the current image (first row is header)
        for index, row in enumerate(self.ws.iter_rows(min_row=2, values_only=False)):
            if row[0].value == self.image_list[self.image_index]:  # Check if the image matches
                # Update the current row with the annotations
                for i, field in enumerate(self.fields):
                    row[i + 1].value = self.annotations[self.image_list[self.image_index]][field]
                self.wb.save(self.xlsx_file)  # Save the workbook
                break

    def add_field(self):
        field_name = simpledialog.askstring("Add Field", "Enter checkbox name:")
        if field_name and field_name not in self.fields:
            self.fields.append(field_name)
            self.update_checkboxes()  # Immediately update checkboxes on screen
            # Re-create annotation file headers if folder is loaded
            if self.folder_path:
                self.create_annotation_file(self.folder_path)

            # Ask for shortcut assignment
            shortcut = simpledialog.askstring("Assign Shortcut", f"Assign a shortcut for '{field_name}':")
            if shortcut and len(shortcut) == 1:
                self.field_shortcuts[field_name] = shortcut

    def toggle_field(self, event):
        field_name = next((name for name, shortcut in self.field_shortcuts.items() if shortcut == event.char), None)
        if field_name:
            var = self.field_vars[field_name]  # Get the IntVar for the field
            var.set(1 - var.get())  # Toggle checkbox state
            self.update_annotation(field_name, var)


if __name__ == "__main__":
    root = tk.Tk()
    app = ImageAnnotator(root)
    root.mainloop()
